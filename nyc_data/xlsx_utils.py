import csv
import json
import re
from pathlib import Path
from typing import NamedTuple, Any, Callable, List, Optional, Set, Union

from django.core.serializers.json import DjangoJSONEncoder
from fuzzywuzzy import process
from openpyxl import load_workbook

from ppe import errors
from ppe.data_mapping.types import DataFile
from ppe.data_mapping.utils import ErrorCollector
from ppe.errors import ColumnNameMismatch


def XLSXDictReader(sheet, header_row):
    rows = sheet.max_row
    cols = sheet.max_column

    def item(i, j):
        return (
            sheet.cell(row=header_row, column=j).value,
            sheet.cell(row=i, column=j).value,
        )

    return (
        dict(item(i, j) for j in range(1, cols + 1))
        for i in range(header_row + 1, rows + 1)
    )


class Mapping(NamedTuple):
    sheet_column_name: str
    obj_column_name: str
    proc: Optional[Callable[[str, ErrorCollector], Any]] = None


def ExactMatch(sheet_name: str):
    def matcher(names):
        if sheet_name in names:
            return sheet_name
        else:
            return None

    return matcher


class RegexMatch:
    def __init__(self, patt):
        self.patt = patt

    def __repr__(self):
        return self.patt

    def __call__(self, names):
        opts = [name for name in names if re.match(self.patt, name)]
        if len(opts) == 1:
            return opts[0]
        if len(opts) > 1:
            raise Exception("Too many matches for sheet pattern")
        return None


class SheetMapping(NamedTuple):
    data_file: DataFile
    sheet_name: Optional[Union[Callable[[List[str]], Optional[str]], str]]
    mappings: Set[Mapping]
    include_raw: bool
    obj_constructor: Optional[Callable[[Any], "ImportedRow"]]
    header_row_idx: int = 1

    def load_data(self, path: Path):
        if self.sheet_name is None:
            try:
                with open(path, encoding="latin-1") as csvfile:
                    text = csvfile.read()
            except Exception as exc:
                raise errors.CsvImportError("Error reading in CSV file") from exc

            return csv.DictReader(text.splitlines())
        else:
            workbook = load_workbook(path, data_only=True)
            actual_sheet = self.can_import(workbook.sheetnames)
            if actual_sheet is None:
                raise Exception(
                    "Tried to import a sheet with a data mapping that does not match"
                )
            return XLSXDictReader(workbook[actual_sheet], self.header_row_idx)

    def can_import(self, sheet_names):
        if isinstance(self.sheet_name, str):
            return ExactMatch(self.sheet_name)(sheet_names)
        elif self.sheet_name is None:
            return None
        else:
            return self.sheet_name(sheet_names)

    def key_columns(self):
        return (mapping.sheet_column_name for mapping in self.mappings)


RAW_DATA = "raw_data"


def guess_mapping(sheet: Path, all_mappings: List[SheetMapping]):
    workbook = None
    if sheet.suffix == ".xlsx":
        workbook = load_workbook(sheet, data_only=True)
        possible_mappings = [
        m for m in all_mappings if m.can_import(workbook.sheetnames)
        ]
        if not possible_mappings:
            known_sheetnames = [m.sheet_name for m in all_mappings]
            matches = [
                (us, process.extractOne(us, known_sheetnames))
                for us in workbook.sheetnames
            ]
            raise errors.SheetNameMismatch(
                workbook.sheetnames, (matches[0][0], matches[0][1][0])
            )
        else:
            df = possible_mappings[0].data_file

            if len([m for m in all_mappings if m.data_file == df]) != len(
                possible_mappings
            ):
                expected = [m.sheet_name for m in all_mappings if m.data_file == df]
                raise errors.PartialFile(
                    expected_sheets=expected, actual_sheets=workbook.sheetnames
                )

    elif sheet.suffix == ".csv":
        possible_mappings = [m for m in all_mappings if m.sheet_name is None]
    else:
        return []
    final_mappings = []

    for mapping in possible_mappings:
        first_row = next(mapping.load_data(sheet))

        col_names = [m.sheet_column_name for m in mapping.mappings]
        if all(col_name in first_row for col_name in col_names):
            final_mappings.append(mapping)
        elif mapping.sheet_name is not None:
            raise ColumnNameMismatch(col_names, first_row.keys())

    if final_mappings:
        return final_mappings


def import_xlsx(
    path: Path,
    sheet_mapping: SheetMapping,
    error_collector: ErrorCollector = lambda: ErrorCollector(),
):
    as_dicts = list(sheet_mapping.load_data(path))

    for row in as_dicts:
        mapped_row = {}
        if all(row.get(col) is None for col in sheet_mapping.key_columns()):
            continue
        for mapping in sheet_mapping.mappings:
            item = row[mapping.sheet_column_name]
            if mapping.proc:
                item = mapping.proc(item, error_collector)
            mapped_row[mapping.obj_column_name] = item

        if sheet_mapping.include_raw:
            # allow serialization of datetimes
            mapped_row[RAW_DATA] = json.dumps(row, cls=DjangoJSONEncoder)
        if sheet_mapping.obj_constructor:
            yield sheet_mapping.obj_constructor(**mapped_row)
        else:
            yield mapped_row
